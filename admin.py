'''This file contains extra feature 
that are supposed to admin should have.'''

import sqlite3
from flask import (
            Blueprint,
            render_template, 
            request, 
            url_for, 
            redirect, 
            session, abort
)
from openpyxl import load_workbook

def get_data(name):
    conn = get_db_connection()
    data = conn.execute("SELECT * FROM userlist WHERE _name=?", (name,)).fetchone()
    conn.close()
    if data is None:
        abort(404)
    return data

def get_db_connection():
    conn = sqlite3.connect("C:\\Users\\jsk\\Desktop\\school project\\code\\code\\database.db")
    conn.row_factory = sqlite3.Row
    return conn

admin_file = Blueprint('admin_file', __name__, template_folder='templates')

@admin_file.route('/<id>/add-student', methods=('GET', 'POST'))
def addstudent(id):
    data = get_data(id)
    if session.get('admin') == data['_name']:
        if request.method == 'POST':
            enrollmentyear = request.form.get('enrollmentyear')
            standard = request.form.get('standard')
            excel_file = request.files['file']

            wb = load_workbook(excel_file)
            ws = wb.active
            
            conn = get_db_connection()
            
            for row in ws.iter_rows(min_row=2):
                conn.execute('INSERT INTO userlist (_name, _password, isAdmin) VALUES(?, ?, ?)', (row[0].value, row[2].value, 0))
                conn.execute("INSERT INTO studentprofile (s_fname, s_mname, s_lname, gender, birthdate, cast_catagory, '10thpr', '10thpercentage', past_school_name, enrollment_year, standard, permenent_id) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, enrollmentyear, standard, "".join([str(row[0].value), str(enrollmentyear)])))
                conn.execute("INSERT INTO student_contact (s_fname, home_address, city, pin_code, home_state, mobile_no, email) VALUES(?, ?, ?, ?, ?, ?, ?)", (row[0].value, row[9].value, row[10].value, row[11].value, row[12].value, row[13].value, row[14].value))

            conn.commit()
            conn.close()
            return redirect(url_for('index'))
        return render_template('add_student.html', list=data)
    else:
        redirect(url_for('login'))

@admin_file.route('/<name>/hall-ticket-admin/')
def hallticket(name):
    data = get_data(name)
    conn = get_db_connection()
    h_data = conn.execute("SELECT * FROM 'Hall Ticket'")
    conn.close()
    if session.get('admin') == data['_name']:
        if request.method == "POST":
            exam_no = request.form.get('examnumber')
            exam_code = request.form.get('examcode')
            exam_type = request.form.get('examtype')
            exam_date = request.form.get('examdate')
            semester = request.form.get('semester')
            exam_time = request.form.get('timeofexam')

            conn = get_db_connection()

            conn.execute("INSERT INTO 'Hall Ticket' ('Exam No', 'Exam Code', 'Exam Type', 'Date', 'Semester', 'Exam Time') VALUES(?, ?, ?, ?, ?)", (exam_no, exam_code, exam_type, exam_date, semester, exam_time ))

            conn.commit()
            conn.close()
        
        return render_template("hallticket_admin.html", list=data, data=h_data)
    else:
        return redirect(url_for('login'))




@admin_file.route('/<name>/fee-voucher-admin/')
def FeeVoucher(name):
    data = get_data(name)
    if(data['_name'] == session.get('admin')):
        return render_template('fee-voucher-admin.html', list=data)
    else:
        abort(404, "Admin fee voucher page does not found")


@admin_file.route('/<name>/examresult-admin/')
def ExamResult(name):
    data = get_data(name)
    if(data['_name'] == session.get('admin')):
        return render_template('examresult-admin.html', list=data)
    else:
        abort(404)

@admin_file.route('/<name>/subjects-admin', methods=['GET', 'POST'])
def Subjects(name):
    data = get_data(name)
    conn = get_db_connection()
    subject = conn.execute("SELECT * FROM 'Subjects'").fetchall()
    if data['_name'] == session.get('admin'):
        if request.method == 'POST':
            excel_file = request.files['file']

            wb = load_workbook(excel_file)
            ws = wb.active
            
            conn = get_db_connection()
            
            for row in ws.iter_rows(min_row=2):
                conn.execute("INSERT INTO 'Subjects' ('Subject Name', Standard, Semester, 'Subject Credit') VALUES(?, ?, ?, ?)", (row[0].value, row[1].value, row[2].value, row[3].value))
            

            conn.commit()
            conn.close()
            return redirect(f'/{name}/home')

        return render_template('subjects_admin.html', subjects=subject, list=data)
    else:
        abort(404)